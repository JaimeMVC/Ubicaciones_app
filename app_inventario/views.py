from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.urls import reverse
from django.contrib import messages
import csv

import pandas as pd
from django.db import transaction

from .models import (
    LocationBase,
    CountSession,
    CountDetail,
    ResultSnapshot,  # aunque casi no lo usemos ahora, lo dejamos
)


# ========= Helpers para importar Excel =========

def _norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = (
        s.replace("á", "a")
         .replace("é", "e")
         .replace("í", "i")
         .replace("ó", "o")
         .replace("ú", "u")
         .replace("ñ", "n")
    )
    for ch in [" ", "\t", "\n", "-", "_", ".", "/"]:
        s = s.replace(ch, "")
    return s


from openpyxl import load_workbook

def _import_df_to_locationbase(file) -> int:
    """
    Importa Excel sin usar pandas.
    Procesa fila por fila en streaming → no consume RAM.
    """

    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    # Leer encabezados
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Normalizar nombres
    def norm(s):
        return (
            s.lower()
            .replace(" ", "")
            .replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u").replace("ñ","n")
            .replace("_","").replace("-","").replace(".","").replace("/","")
        )

    norm_headers = [norm(h) for h in headers]

    def pick(names):
        for name in names:
            if name in norm_headers:
                return norm_headers.index(name)
        return None

    idx_pn = pick(["pn","partnumber","material","codigo","codigomaterial"])
    idx_ubi = pick(["ubicacion","ubicaciones","location"])
    idx_des = pick(["descripcion","description","desc"])

    if idx_pn is None or idx_ubi is None:
        raise ValueError("No se encontraron columnas PN y Ubicacion")

    # Marcar todo como inactivo
    LocationBase.objects.update(activo=False)

    count = 0
    rows_to_create = []

    for row in ws.iter_rows(min_row=2):
        pn = row[idx_pn].value if idx_pn < len(row) else None
        ub = row[idx_ubi].value if idx_ubi < len(row) else None
        desc = row[idx_des].value if idx_des is not None and idx_des < len(row) else ""

        if not pn or not ub:
            continue

        pn = str(pn).strip()
        ub = str(ub).strip()
        desc = str(desc).strip() if desc else ""

        if not pn or not ub:
            continue

        # update_or_create no consume memoria
        LocationBase.objects.update_or_create(
            pn=pn,
            ubicacion=ub,
            defaults={"descripcion": desc, "activo": True}
        )

        count += 1

    return count



# ========= Vistas principales =========

def cargar_excel(request):
    """Subir Excel desde la web y actualizar LocationBase."""
    if request.method == "POST" and request.FILES.get("archivo"):
        try:
            n = _import_df_to_locationbase(request.FILES["archivo"])

            messages.success(request, f"Archivo importado correctamente. Filas procesadas: {n}.")
            return redirect("buscar_material")
        except Exception as e:
            messages.error(request, f"Error al importar: {e}")
            return redirect("cargar_excel")

    return render(request, "app_inventario/cargar_excel.html")


def buscar_material(request):
    query = request.GET.get("q", "").strip()
    materiales = []

    if query:
        materiales = (
            LocationBase.objects
            .filter(pn__icontains=query, activo=True)
            .values("pn")
            .distinct()
            .order_by("pn")
        )

    return render(request, "app_inventario/buscar_material.html", {
        "query": query,
        "materiales": materiales,
    })


def listado_ubicaciones(request, pn):
    """
    - Si es POST: crea una nueva CountSession para ese PN.
    - Si es GET:
        - Sin ?session= → muestra formulario para crear sesión y listado de sesiones previas.
        - Con ?session=ID → muestra checklist para esa sesión.
    """
    # Crear nueva sesión
    if request.method == "POST":
        operador = request.POST.get("operador", "").strip()
        comentario = request.POST.get("comentario", "").strip()
        if not operador:
            messages.error(request, "Debe ingresar el nombre del operador.")
        else:
            session = CountSession.objects.create(
                pn=pn,
                operador=operador,
                comentario=comentario,
            )
            return redirect(f"{reverse('listado_ubicaciones', args=[pn])}?session={session.id}")

    session_id = request.GET.get("session")
    sesiones_pn = CountSession.objects.filter(pn=pn).order_by("-creado_en")

    session = None
    rows = []
    total = revisadas = 0

    if session_id:
        session = get_object_or_404(CountSession, id=session_id, pn=pn)
        base_ubics = list(LocationBase.objects.filter(pn=pn, activo=True).order_by("ubicacion"))
        detalles = CountDetail.objects.filter(session=session)
        detalles_by_base = {d.base_id: d for d in detalles}

        for b in base_ubics:
            detalle = detalles_by_base.get(b.id)
            rows.append({"base": b, "detalle": detalle})
            if detalle and detalle.revisado:
                revisadas += 1

        total = len(base_ubics)

    porcentaje = round(revisadas / total * 100, 1) if total else 0.0

    return render(request, "app_inventario/listado_ubicaciones.html", {
        "pn": pn,
        "session": session,
        "sesiones_pn": sesiones_pn,
        "rows": rows,
        "total": total,
        "revisadas": revisadas,
        "porcentaje": porcentaje,
    })


def toggle_check(request):
    """Marca/desmarca una ubicación dentro de una sesión."""
    if request.method == "POST":
        session_id = request.POST.get("session_id")
        base_id = request.POST.get("base_id")
        checked = request.POST.get("checked") == "true"

        session = get_object_or_404(CountSession, id=session_id)
        base = get_object_or_404(LocationBase, id=base_id)

        detail, _ = CountDetail.objects.get_or_create(session=session, base=base)
        detail.revisado = checked
        detail.fecha_revision = timezone.now() if checked else None
        detail.save()

        return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)


def historial_pn(request, pn):
    """Historial de sesiones para un PN, con avance calculado."""
    sesiones = CountSession.objects.filter(pn=pn).order_by("-creado_en")

    data = []
    for s in sesiones:
        base_ubics = LocationBase.objects.filter(pn=pn, activo=True)
        total = base_ubics.count()
        revisadas = CountDetail.objects.filter(session=s, revisado=True).count()
        porcentaje = round(revisadas / total * 100, 1) if total else 0.0
        data.append({
            "session": s,
            "total": total,
            "revisadas": revisadas,
            "porcentaje": porcentaje,
        })

    return render(request, "app_inventario/historial_pn.html", {
        "pn": pn,
        "data": data,
    })


def informe_sesion(request, session_id):
    """Informe detallado de una sesión de conteo."""
    session = get_object_or_404(CountSession, id=session_id)
    pn = session.pn

    base_ubics = list(LocationBase.objects.filter(pn=pn, activo=True).order_by("ubicacion"))
    detalles = CountDetail.objects.filter(session=session)
    detalles_by_base = {d.base_id: d for d in detalles}

    rows = []
    total = len(base_ubics)
    revisadas = 0
    for b in base_ubics:
        detalle = detalles_by_base.get(b.id)
        if detalle and detalle.revisado:
            revisadas += 1
        rows.append({"base": b, "detalle": detalle})

    porcentaje = round(revisadas / total * 100, 1) if total else 0.0

    return render(request, "app_inventario/informe_sesion.html", {
        "session": session,
        "pn": pn,
        "rows": rows,
        "total": total,
        "revisadas": revisadas,
        "porcentaje": porcentaje,
    })


def exportar_sesion_csv(request, session_id):
    """Exporta a CSV una sesión de conteo."""
    session = get_object_or_404(CountSession, id=session_id)
    pn = session.pn

    base_ubics = list(LocationBase.objects.filter(pn=pn, activo=True).order_by("ubicacion"))
    detalles = CountDetail.objects.filter(session=session)
    detalles_by_base = {d.base_id: d for d in detalles}

    # Preparamos respuesta CSV
    filename = f"avance_{pn}_sesion_{session.id}.csv"
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        "PN",
        "Operador",
        "Fecha sesión",
        "Ubicación",
        "Descripción",
        "Revisado",
        "Fecha revisión",
        "Comentario sesión",
    ])

    for b in base_ubics:
        detalle = detalles_by_base.get(b.id)
        writer.writerow([
            pn,
            session.operador,
            session.creado_en.strftime("%Y-%m-%d %H:%M"),
            b.ubicacion,
            b.descripcion,
            "SI" if (detalle and detalle.revisado) else "NO",
            detalle.fecha_revision.strftime("%Y-%m-%d %H:%M") if (detalle and detalle.fecha_revision) else "",
            session.comentario or "",
        ])

    return response
